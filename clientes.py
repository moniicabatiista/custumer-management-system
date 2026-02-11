import sqlite3
import pandas as pd
from openpyxl import load_workbook
conexao = sqlite3.connect("clientes.db")
cursor = conexao.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS clientes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    nome TEXT,
    email TEXT
)
""")
def atualizar_excel():
    try:
        df = pd.read_sql_query("SELECT * FROM clientes", conexao)

        # escreve os dados começando na linha 3
        df.to_excel("clientes.xlsx", index=False, startrow=2)

        wb = load_workbook("clientes.xlsx")
        ws = wb.active

        # título
        ws["A1"] = "CADASTRO DE CLIENTES"
        ws["A1"].font = ws["A1"].font.copy(bold=True)
        ws.merge_cells("A1:C1")

        wb.save("clientes.xlsx")

    except PermissionError:
        print("⚠️ Feche o arquivo clientes.xlsx antes de atualizar.")
    
def cadastrar_cliente():
    nome = input("Nome Completo: ")
    email = input("Email: ")
    cursor.execute(
        "INSERT INTO clientes (nome, email) VALUES (?, ?)",
        (nome, email)
    )
    conexao.commit()
    atualizar_excel()

def listar_clientes():
    clientes = cursor.execute("SELECT * FROM clientes").fetchall()
    for cliente in clientes:
        print(cliente)

def atualizar_cliente():
    id_cliente = input("ID do cliente: ")
    novo_email = input("Novo email: ")
    cursor.execute(
        "UPDATE clientes SET email = ? WHERE id = ?",
        (novo_email, id_cliente)
    )
    conexao.commit()
    atualizar_excel()

def deletar_cliente():
    id_cliente = int(input("ID do cliente: "))
    cursor.execute("DELETE FROM clientes WHERE id = ?", (id_cliente,))
    conexao.commit()
    atualizar_excel()

while True:
    print("1 - Cadastrar | 2 - Listar | 3 - Atualizar | 4 - Deletar | 0 - Sair")
    opcao = input("Escolha: ")

    if opcao == "1":
        cadastrar_cliente()
    elif opcao == "2":
        listar_clientes()
    elif opcao == "3":
        atualizar_cliente()
    elif opcao == "4":
        deletar_cliente()
    elif opcao == "0":
        break

conexao.close()
